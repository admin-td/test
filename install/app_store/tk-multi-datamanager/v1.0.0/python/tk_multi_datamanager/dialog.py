# Copyright (c) 2017 Shotgun Software Inc.
#
# CONFIDENTIAL AND PROPRIETARY
#
# This work is provided "AS IS" and subject to the Shotgun Pipeline Toolkit
# Source Code License included in this distribution package. See LICENSE.
# By accessing, using, copying or modifying this work you indicate your
# agreement to the Shotgun Pipeline Toolkit Source Code License. All rights
# not expressly granted therein are reserved by Shotgun Software Inc.
import json
import os
import shutil
import time
import traceback
import re
import math

import sgtk
from sgtk.platform.qt import QtCore, QtGui
from tank_vendor import six
import sys
import subprocess

from .api import PublishManager, PublishItem, PublishTask
from .ui.dialog import Ui_Dialog
from .progress import ProgressHandler
from .summary_overlay import SummaryOverlay
from .publish_tree_widget import TreeNodeItem, TreeNodeTask, TopLevelTreeNodeItem

browse_icon_path = (r"X:\Inhouse\tool\icon\browse.png")
scan_icon_path = (r"X:\Inhouse\tool\icon\browse_menu.png")
copy_icon_path = (r"X:\Inhouse\tool\icon\image_sequence.png")
venv_path = r'X:\Inhouse\Python\.venv\Lib\site-packages'
sys.path.append(venv_path)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from pydpx_meta import DpxHeaderEx
from pymediainfo import MediaInfo
from OpenEXR import InputFile
import pandas as pd

# import frameworks
settings = sgtk.platform.import_framework("tk-framework-shotgunutils", "settings")
help_screen = sgtk.platform.import_framework("tk-framework-qtwidgets", "help_screen")
task_manager = sgtk.platform.import_framework(
    "tk-framework-shotgunutils", "task_manager"
)
shotgun_model = sgtk.platform.import_framework(
    "tk-framework-shotgunutils", "shotgun_model"
)
shotgun_globals = sgtk.platform.import_framework(
    "tk-framework-shotgunutils", "shotgun_globals"
)

logger = sgtk.platform.get_logger(__name__)


class CheckableItem(QtGui.QStandardItem):
    def __init__(self, file_info_dict='', save_func=True):

        super().__init__(file_info_dict['path'])
        parts = file_info_dict['path'].split('/')
        date_name = parts[-2]

        # Retrieve checked items
        self.checked_items = CheckableItem.fetch_checked_items(date_name)

        if save_func:
            self.setFlags(QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)

            if not self.checked_items:
                self.setData(QtCore.Qt.Checked, QtCore.Qt.CheckStateRole)
            else:
                # Check if the current item should be checked
                for item in self.checked_items:
                    if item['Shot_Code'] + '_' + item['Type'] + '_' + item['Version'] in file_info_dict['path']:
                        self.setData(QtCore.Qt.Unchecked, QtCore.Qt.CheckStateRole)
                        break
                    else:
                        self.setData(QtCore.Qt.Checked, QtCore.Qt.CheckStateRole)

    @staticmethod
    def fetch_checked_items(date_name):
        """
        Fetch checked items from the latest Excel file.
        """

        # Get the latest Excel file path
        bundle = sgtk.platform.current_bundle()
        default_directory = "X:/ShotGrid_Test_jw/Project/" + bundle.context.project.get("name", "Undefined") + '/scan' + '/' + date_name
        excel_file_path = CheckableItem._get_latest_excel_file_path(default_directory)

        if not excel_file_path:
            return None

        # Read the Excel file
        df = pd.read_excel(excel_file_path)

        # Filter checked and unchecked rows
        checked_df = df[df['Check'] == 'checked']

        # Create a list of dictionaries containing index and row data for checked items
        checked_items_list = [{'index': idx, **row} for idx, row in checked_df.iterrows()]

        return checked_items_list

    @staticmethod
    def _get_latest_excel_file_path(directory):
        """
        Get the latest Excel file from the specified directory.

        Args:
            directory (str): The directory to search for Excel files.

        Returns:
            str: The path to the latest Excel file.
        """
        base_name = "scan_list_"
        ext = ".xlsx"

        # Find all files in the directory that match the base name and extension
        existing_files = [f for f in os.listdir(directory) if f.startswith(base_name) and f.endswith(ext)]

        if not existing_files:
            return None

        # Extract numbers from filenames and find the highest number
        numbers = [int(re.findall(r'\d+', f[len(base_name):])[0]) for f in existing_files if
                   re.findall(r'\d+', f[len(base_name):])]
        max_number = max(numbers) if numbers else 1
        next_file_name = f"{base_name}{max_number:02d}{ext}"

        return os.path.join(directory, next_file_name)


class AppDialog(QtGui.QWidget):
    """
    Main dialog window for the App
    """

    # main drag and drop areas
    (DRAG_SCREEN, PUBLISH_SCREEN) = range(2)

    # details ui panes
    (
        ITEM_DETAILS,
        TASK_DETAILS,
        PLEASE_SELECT_DETAILS,
        MULTI_EDIT_NOT_SUPPORTED,
    ) = range(4)

    def __init__(self, parent=None):
        """
        :param parent: The parent QWidget for this control
        """
        QtGui.QWidget.__init__(self, parent)

        # create a settings manager where we can pull and push prefs later
        # prefs in this manager are shared
        self._settings_manager = settings.UserSettings(sgtk.platform.current_bundle())

        # create a background task manager
        self._task_manager = task_manager.BackgroundTaskManager(
            self, start_processing=True, max_threads=2
        )

        # register the data fetcher with the global schema manager
        shotgun_globals.register_bg_task_manager(self._task_manager)

        self._bundle = sgtk.platform.current_bundle()
        self._validation_run = False
        self._default_directory_path = "X:/ShotGrid_Test_jw/Project/" + self._bundle.context.project.get("name", "Undefined") + '/scan'
        self._default_config_path = self._bundle.sgtk.configuration_descriptor.get_path() + r'\install\app_store\tk-multi-datamanager\v1.0.0\python\tk_multi_datamanager\batch_render.py'
        self._paths = None
        self._selected_color = 'ACES - ACES2065-1'

        # set up the UI
        self.ui = Ui_Dialog()
        self.ui.setupUi(self)
        self.ui.context_widget.set_up(self._task_manager)

        # set up the Model
        self.model = QtGui.QStandardItemModel()
        self.ui.listView.setModel(self.model)

        self.ui.browse_button.clicked.connect(lambda: self._on_browse(folders=True))
        self.ui.excel_save_button.clicked.connect(self._save_checked_items_to_excel)
        self.ui.check_button.clicked.connect(self._check_all_items)
        self.ui.uncheck_button.clicked.connect(self._uncheck_all_items)
        self.ui.excel_open_button.clicked.connect(self._on_excel_browse)
        self.ui.publish_button.clicked.connect(self._on_drop)
        self.ui.scan_button.clicked.connect(self._scan_listener)
        self.ui.copy_button.clicked.connect(self._copy_listener)

        self.ui.browse_button.setIcon(QtGui.QIcon(browse_icon_path))
        self.ui.scan_button.setIcon(QtGui.QIcon(scan_icon_path))
        self.ui.copy_button.setIcon(QtGui.QIcon(copy_icon_path))

        self.ui.comboBox.addItem("ACES - ACEScg")
        self.ui.comboBox.addItem("AlexaV3LogC")
        self.ui.comboBox.addItem("AlexaV4LogC")
        self.ui.comboBox.addItem("rec709")
        self.ui.comboBox.addItem("sRGB")
        self.ui.comboBox.addItem("legacy")
        self.ui.comboBox.currentIndexChanged.connect(self._on_combobox_changed)

        # only allow entities that can be linked to PublishedFile entities
        self.ui.context_widget.restrict_entity_types_by_link("PublishedFile", "entity")

        # tooltips for the task and link inputs
        self.ui.context_widget.set_task_tooltip(
            "<p>The task that the selected item will be associated with "
            "in Flow Production Tracking after publishing. It is recommended "
            "to always fill out the Task field when publishing. The menu button "
            "to the right will provide suggestions for Tasks to publish "
            "including the Tasks assigned to you, recently used Tasks, "
            "and Tasks related to the entity Link populated in the field "
            "below.</p>"
        )
        self.ui.context_widget.set_link_tooltip(
            "<p>The entity that the selected item will be associated with "
            "in Flow Production Tracking after publishing. By selecting a "
            "Task in the field above, the Link will automatically be populated. "
            "It is recommended that you always populate the Task field when "
            "publishing. The Task menu above will display any tasks associated "
            "with the entity populated in this field.</p>"
        )

        self.ui.context_widget.context_changed.connect(self._on_item_context_change)

        # make sure the splitter expands the detail area only
        self.ui.splitter.setStretchFactor(0, 0)
        self.ui.splitter.setStretchFactor(1, 1)

        # give tree view 360 width, rest to details pane
        # note: value of second option does not seem to
        # matter (as long as it's there)
        self.ui.splitter.setSizes([360, 100])

        # drag and drop
        self.ui.frame.something_dropped.connect(self._on_drop)

        self.ui.items_tree.tree_reordered.connect(self._synchronize_tree)

        # hide the drag screen progress button by default
        # buttons
        self.ui.publish.clicked.connect(self.do_publish)
        self.ui.close.clicked.connect(self.close)
        self.ui.close.hide()
        # do_validate is invoked using a lambda function because it receives custom parameters
        # https://eli.thegreenplace.net/2011/04/25/passing-extra-arguments-to-pyqt-slot/
        self.ui.validate.clicked.connect(lambda: self.do_validate())

        # overlay
        self._overlay = SummaryOverlay(self.ui.main_frame)
        self._overlay.publish_again_clicked.connect(self._publish_again_clicked)

        # settings
        self.ui.items_tree.status_clicked.connect(self._on_publish_status_clicked)

        # Description
        self.ui.item_comments.textChanged.connect(self._on_item_comment_change)
        self.ui.item_inherited_item_label.linkActivated.connect(
            self._on_description_inherited_link_activated
        )

        # selection in tree view
        self.ui.items_tree.itemSelectionChanged.connect(
            self._update_details_from_selection
        )

        # clicking in the tree view
        self.ui.items_tree.checked.connect(self._update_details_from_selection)

        # thumbnails
        self.ui.item_thumbnail.thumbnail_changed.connect(self._update_item_thumbnail)

        # tool buttons
        self.ui.delete_items.clicked.connect(self._delete_selected)
        self.ui.expand_all.clicked.connect(lambda: self._set_tree_items_expanded(True))
        self.ui.collapse_all.clicked.connect(
            lambda: self._set_tree_items_expanded(False)
        )
        self.ui.refresh.clicked.connect(self._full_rebuild)

        # stop processing logic
        # hide the stop processing button by default
        self.ui.stop_processing.hide()
        self._stop_processing_flagged = False
        self.ui.stop_processing.clicked.connect(self._trigger_stop_processing)

        # help button
        help_url = self._bundle.get_setting("help_url")
        if help_url:
            # connect the help button to the help url provided in the settings
            self.ui.help.clicked.connect(lambda: self._open_url(help_url))
        else:
            # no url. hide the button!
            self.ui.help.hide()

        # browse file action
        self._browse_file_action = QtGui.QAction(self)
        self._browse_file_action.setText("Browse files to publish")
        self._browse_file_action.setIcon(QtGui.QIcon(":/tk_multi_publish2/file.png"))
        self._browse_file_action.triggered.connect(
            lambda: self._on_browse(folders=False)
        )

        # browse folder action
        self._browse_folder_action = QtGui.QAction(self)
        self._browse_folder_action.setText("Browse folders to publish image sequences")
        self._browse_folder_action.setIcon(
            QtGui.QIcon(":/tk_multi_publish2/image_sequence.png")
        )
        self._browse_folder_action.triggered.connect(
            lambda: self._on_browse(folders=True)
        )

        # browse menu
        self._browse_menu = QtGui.QMenu(self)
        self._browse_menu.addAction(self._browse_file_action)
        self._browse_menu.addAction(self._browse_folder_action)

        # browse tool button
        self.ui.browse.clicked.connect(self._on_browse)
        self.ui.browse.setMenu(self._browse_menu)
        self.ui.browse.setPopupMode(QtGui.QToolButton.DelayedPopup)

        # drop area browse button. Note, not using the actions created above
        # because making the buttons look right when they're using he action's
        # text/icon proved difficult. Instead, the button text/icon are defined
        # in the designer file. So as a note, if you want to change the text or
        # icon, you'll need to do it above and in designer.

        # currently displayed item
        self._current_item = None

        # Currently selected tasks. If a selection is created in the GUI that
        # contains multiple task types or even other tree item types, then,
        # _current_tasks will be set to an empty selection, regardless of the
        # number of the items actually selected in the UI.
        self._current_tasks = _TaskSelection()

        self._summary_comment = ""
        self._scan_path = None

        # set up progress reporting
        self._progress_handler = ProgressHandler(
            self.ui.progress_status_icon, self.ui.progress_message, self.ui.progress_bar
        )

        # link the summary overlay status button with the log window
        self._overlay.info_clicked.connect(
            self._progress_handler._progress_details.toggle
        )

        # hide settings for now
        self.ui.item_settings_label.hide()
        self.ui.item_settings.hide()

        # create a publish manager
        self._publish_manager = PublishManager(self._progress_handler.logger)
        self.ui.items_tree.set_publish_manager(self._publish_manager)

        # this is the pixmap in the summary thumbnail
        self._summary_thumbnail = None

        # set publish button text
        self._display_action_name = self._bundle.get_setting("display_action_name")
        self.ui.publish.setText(self._display_action_name)

        # Special UI tweaks based on the 'manual_load_enabled' property
        #
        # Hide the tiny label at bottom of the tree view and
        # the browse button in the button container
        self.ui.text_below_item_tree.setVisible(self.manual_load_enabled)
        self.ui.browse.setVisible(self.manual_load_enabled)

        # run collections
        self._full_rebuild()

    def _on_combobox_changed(self):
        self._selected_color = self.ui.comboBox.currentText()
        logger.info(f"Selected Color: {self._selected_color}")

    @property
    def manual_load_enabled(self):
        """Returns whether user is allowed to load file to the UI"""
        return self._bundle.get_setting("enable_manual_load")

    def keyPressEvent(self, event):
        """
        Qt Keypress event
        """
        # if our log details ui is showing and escape
        # is pressed, capture it and hide the log details ui
        if (
            self._progress_handler.is_showing_details()
            and event.key() == QtCore.Qt.Key_Escape
        ):
            # hide log window
            self._progress_handler.hide_details()

        else:
            super(AppDialog, self).keyPressEvent(event)

    def closeEvent(self, event):
        """
        Executed when the main dialog is closed.
        All worker threads and other things which need a proper shutdown
        need to be called here.
        """
        logger.debug("CloseEvent Received. Begin shutting down UI.")

        # register the data fetcher with the global schema manager
        shotgun_globals.unregister_bg_task_manager(self._task_manager)

        # deallocate loggers
        self._progress_handler.shut_down()

        try:
            # shut down main threadpool
            self._task_manager.shut_down()
        except Exception:
            logger.exception(
                "Error running Flow Production Tracking Panel App closeEvent()"
            )

        # ensure the context widget's recent contexts are saved
        self.ui.context_widget.save_recent_contexts()

    def _update_details_from_selection(self):
        """
        Makes sure that the right hand side details section reflects the selected item in the left
        hand side tree.
        """

        # look at how many items are checked
        checked_top_items = 0
        for context_index in range(self.ui.items_tree.topLevelItemCount()):
            context_item = self.ui.items_tree.topLevelItem(context_index)
            for child_index in range(context_item.childCount()):
                child_item = context_item.child(child_index)
                if child_item.checked:
                    checked_top_items += 1

        if checked_top_items == 0:
            # disable buttons
            self.ui.publish.setEnabled(False)
            self.ui.validate.setEnabled(False)
        else:
            self.ui.publish.setEnabled(True)
            self.ui.validate.setEnabled(True)

        # now look at selection
        items = self.ui.items_tree.selectedItems()

        if self._is_task_selection_homogeneous(items):
            # We should update the tasks details ui.
            self._current_item = None
            publish_tasks = _TaskSelection(
                [item.get_publish_instance() for item in items]
            )
            self._update_task_details_ui(publish_tasks)
        elif len(items) != 1:
            # Otherwise we can't show items from a multi-selection, so inform the user.
            self._current_item = None
            self._update_task_details_ui()
            # show overlay with 'please select single item'
            self.ui.details_stack.setCurrentIndex(self.PLEASE_SELECT_DETAILS)
        else:
            # 1 item selected
            tree_item = items[0]
            publish_object = tree_item.get_publish_instance()
            if isinstance(publish_object, PublishItem):
                self._update_task_details_ui()
                self._create_item_details(tree_item)
            elif publish_object is None:
                self._update_task_details_ui()
                # top node summary
                self._create_master_summary_details()

        # Make the task validation if the setting `task_required` exists and it's True
        self._validate_task_required()

    def _is_task_selection_homogeneous(self, items):
        """
        Indicates if a selection is made up only of tasks and they are all of the same item.

        :param items: List of tree node items.

        :returns: ``True`` is the selection only contains tasks, ``False`` otherwise.
        """
        # If the list is empty, we don't have a task selection.
        if len(items) == 0:
            return False

        # Grab the first item in the list, which we will use to compare to every other item. If
        # all items end up being the same type as the first one, then we have a homogeneous list
        # of tasks.
        first_task = items[0].get_publish_instance()

        for item in items:

            publish_instance = item.get_publish_instance()
            # User has mixed different types of publish instances, it's not just a task list.
            if not isinstance(publish_instance, PublishTask):
                return False

            # There's a task that's not of the same type as the others, we're done!
            if not first_task.is_same_task_type(publish_instance):
                return False

        return True

    def _update_task_details_ui(self, new_task_selection=None):
        """
        Updates the plugin UI widget.

        This method should be called if everything is of the same type OR if the selection is
        empty.

        :param new_task_selection: A :class:`TaskSelection` containing the current UI selection.
        """
        new_task_selection = new_task_selection or _TaskSelection()

        # Nothing changed, so do nothing.
        if self._current_tasks == new_task_selection:
            return

        # We're changing task, so we need to backup the current settings.
        if self._current_tasks:
            logger.debug("Saving settings...")
            self._pull_settings_from_ui(self._current_tasks)

        # If we're moving to a task that doesn't have a custom UI, clear everything.
        if not new_task_selection:
            # Note: At this point we don't really care if current task actually had a UI, we can
            # certainly tear down an empty widget.
            logger.debug("The ui is going to change, so clear the current one.")
            self.ui.task_settings.widget = None
            self._current_tasks = new_task_selection
            return

        # A task was picked, so make sure our page is in foreground.
        self.ui.details_stack.setCurrentIndex(self.TASK_DETAILS)

        # set the header for the task plugin
        self.ui.task_icon.setPixmap(new_task_selection.plugin.icon)
        self.ui.task_name.setText(new_task_selection.plugin.name)

        # Now figure out if we need to create/replace the widgets.
        if (
                # If we had a selection before
                self._current_tasks
                and
                # and it was of the same type as the new one.
                self._current_tasks.is_same_task_type(new_task_selection)
        ):
            logger.debug("Reusing custom ui from %s.", new_task_selection.plugin)
        else:
            logger.debug("Building a custom ui for %s.", new_task_selection.plugin)
            widget = new_task_selection.plugin.run_create_settings_widget(
                self.ui.task_settings_parent, new_task_selection.get_task_items()
            )
            self.ui.task_settings.widget = widget

        # Update the UI with the settings from the current plugin.
        if self._push_settings_into_ui(new_task_selection):
            # Alright, we're ready to deal with the new task.
            self._current_tasks = new_task_selection
        else:
            self._current_tasks = _TaskSelection()

    def _pull_settings_from_ui(self, selected_tasks):
        """
        Retrieves settings from the UI and updates the task's settings.

        :param selected_tasks: A :class:`TaskSelection` of tasks to update based
            on the values edited in the UI.
        """
        if selected_tasks.has_custom_ui:
            widget = self.ui.task_settings.widget
            settings = self._current_tasks.get_settings(widget)
        else:
            # TODO: Implement getting the settings from the generic UI, if we ever implement one.
            settings = {}

        # Update the values in all the tasks.
        for task in selected_tasks:
            # The settings returned by the UI are actual values, not Settings objects, so apply each
            # value returned on the appropriate settings object.
            for k, v in settings.items():
                task.settings[k].value = v

    def _push_settings_into_ui(self, selected_tasks):
        """
        Takes the settings from this task and pushes its values into the UI.

        :param selected_tasks: A :class:`TaskSelection` of tasks to update based
            on the values edited in the UI.
        """
        # The run_get_ui_settings expects a dictionary of the actual values, not Setting objects, so
        # translate the dictionary.
        tasks_settings = []
        for task in selected_tasks:
            settings_dict = {}
            for k, v in task.settings.items():
                settings_dict[k] = v.value
            tasks_settings.append(settings_dict)

        if selected_tasks.has_custom_ui:
            try:
                selected_tasks.set_settings(
                    self.ui.task_settings.widget, tasks_settings
                )
            except NotImplementedError:
                self.ui.details_stack.setCurrentIndex(self.MULTI_EDIT_NOT_SUPPORTED)
                return False
        else:
            # TODO: Implement setting the settings into the generic UI.
            pass
        return True

    def _on_publish_status_clicked(self, task_or_item):
        """
        Triggered when someone clicks the status icon in the tree
        """
        # make sure the progress widget is shown
        self._progress_handler.select_last_message(task_or_item)

    def _on_item_comment_change(self):
        """
        Callback when someone types in the
        publish comments box in the overview details pane
        """
        tree_widget = self.ui.items_tree
        comments = self.ui.item_comments.toPlainText()
        # if this is the summary description...
        if self._current_item is None:
            if self._summary_comment != comments:
                self._summary_comment = comments

                for node_item in tree_widget.root_items():
                    if (
                        isinstance(node_item, TreeNodeItem)
                        and node_item.inherit_description is True
                    ):
                        # This will recursively set all child items that inherit the description.
                        node_item.set_description(comments)

        # the "else" below means if this is a publish item
        else:
            # Define a new variable as we might redefine the comments if the
            # comments are empty we will get the inherited comments.
            description = comments
            # To make the comparison fair, treat a None type description as an empty string.
            item_desc = (
                self._current_item.description
                if self._current_item.description is not None
                else ""
            )
            if item_desc != description:
                # Whilst you can select more than one item, the item description box is not set to show
                # so there can be only one selected item when changing the description.
                node_item = tree_widget.selectedItems()[0]

                if comments == "":
                    # The description has been cleared from the box, so we should retrieve the inherited
                    # description.
                    node_item.inherit_description = True
                    description = self._find_inherited_description(node_item)
                else:
                    # The user has entered the description on this item so we no longer want to
                    # inherit (if it was before).
                    node_item.inherit_description = False

                # This will set all child items that inherit descriptions to the same description.
                node_item.set_description(description)
                self._set_description_inheritance_ui(node_item)

    def _on_description_inherited_link_activated(self, _link):
        """
        When the user activates the link in the inherited from label
        this method will be called. The method selects the item in the tree
        that the current item inherits its description from.
        :param link: The activated URL, we don't need to use this.
        """
        selected_item = self.ui.items_tree.selectedItems()[0]
        item = self._find_inherited_description_item(selected_item)
        if item:
            item.setSelected(True)
        else:
            self.ui.items_tree.summary_node.setSelected(True)
        # Now deselect the item that was selected before clicking the link.
        selected_item.setSelected(False)

    def _update_item_thumbnail(self, pixmap):
        """
        Update the currently selected item with the given
        thumbnail pixmap
        """
        if not self._current_item:
            # this is the summary item
            self._summary_thumbnail = pixmap
            if pixmap:
                # update all items with the summary thumbnail
                for top_level_item in self._publish_manager.tree.root_item.children:
                    top_level_item.thumbnail = self._summary_thumbnail
                    top_level_item.thumbnail_explicit = False

                    # propagate the thumbnail to all descendant items
                    for item in top_level_item.descendants:
                        item.thumbnail = self._summary_thumbnail
                        item.thumbnail_explicit = False
        else:
            self._current_item.thumbnail = pixmap
            # specify that the new thumbnail overrides the one inherited from
            # summary
            self._current_item.thumbnail_explicit = True

    def _create_item_details(self, tree_item):
        """
        Render details pane for a given item
        """
        item = tree_item.get_publish_instance()

        self._current_item = item
        self.ui.details_stack.setCurrentIndex(self.ITEM_DETAILS)
        self.ui.item_icon.setPixmap(item.icon)

        self.ui.item_name.setText(item.name)
        self.ui.item_type.setText(item.type_display)

        # check the state of screenshot
        if item.thumbnail_enabled:
            # display and make thumbnail editable
            self.ui.item_thumbnail_label.show()
            self.ui.item_thumbnail.show()
            self.ui.item_thumbnail.setEnabled(True)

        elif not item.thumbnail_enabled and item.thumbnail:
            # show thumbnail but disabled
            self.ui.item_thumbnail_label.show()
            self.ui.item_thumbnail.show()
            self.ui.item_thumbnail.setEnabled(False)

        # else:
        #     # hide thumbnail
        #     self.ui.item_thumbnail_label.hide()
        #     self.ui.item_thumbnail.hide()

        self.ui.item_description_label.setText("Description")

        # Sets up the UI around the description based on the inheritance state
        self._set_description_inheritance_ui(tree_item)

        if tree_item.inherit_description:
            self.ui.item_comments.setText("")
            self.ui.item_comments.setPlaceholderText(item.description)
        else:
            # We are not inheriting the description so we should set the
            # text box to display the item's description, but we'll also look
            # up what the inherited description would be in case the user clears
            # the box, and it displays the placeholder text.
            self.ui.item_comments.setText(item.description)

            inherited_desc = self._find_inherited_description(tree_item)
            self.ui.item_comments.setPlaceholderText(inherited_desc)

        # Hides the multiple values overlay, as it is only for the summary item.
        self.ui.item_comments._show_multiple_values = False

        # if summary thumbnail is defined, item thumbnail should inherit it
        # unless item thumbnail was set after summary thumbnail
        if self._summary_thumbnail and not item.thumbnail_explicit:
            item.thumbnail = self._summary_thumbnail

        self.ui.item_thumbnail._set_multiple_values_indicator(False)
        self.ui.item_thumbnail.set_thumbnail(item.thumbnail)

        # Items with default thumbnails should still be able to have override thumbnails set by the user
        self.ui.item_thumbnail.setEnabled(True)

        if item.parent.is_root:
            self.ui.context_widget.show()

            if item.context_change_allowed:
                self.ui.context_widget.enable_editing(
                    True, "<p>Task and Entity Link to apply to the selected item:</p>"
                )
            else:
                self.ui.context_widget.enable_editing(
                    False,
                    "<p>Context changing has been disabled for this item. "
                    "It will be associated with "
                    "<strong><a style='color:#C8C8C8; text-decoration:none' "
                    "href='%s'>%s</a></strong></p>"
                    % (item.context.shotgun_url, item.context),
                )

            # set the context
            self.ui.context_widget.set_context(item.context)
        else:
            self.ui.context_widget.hide()

        # create summary
        self.ui.item_summary_label.show()
        summary = tree_item.create_summary()
        # generate a summary

        if len(summary) == 0:
            summary_text = "No items to process."

        else:
            summary_text = "<p>The following items will be processed:</p>"
            summary_text += "".join(["<p>%s</p>" % line for line in summary])

        self.ui.item_summary.setText(summary_text)

        # skip settings for now
        ## render settings
        # self.ui.item_settings.set_static_data(
        #    [(p, item.properties[p]) for p in item.properties]
        # )

    def _create_master_summary_details(self):
        """
        Render the master summary representation
        """
        self._current_item = None
        self.ui.details_stack.setCurrentIndex(self.ITEM_DETAILS)

        display_name = self._bundle.get_setting("display_name")
        self.ui.item_name.setText("%s Summary" % display_name)
        self.ui.item_icon.setPixmap(QtGui.QPixmap(":/tk_multi_publish2/icon_256.png"))

        self.ui.item_thumbnail_label.show()
        self.ui.item_thumbnail.show()

        thumbnail_has_multiple_values = False
        description_had_multiple_values = False
        for top_level_item in self._publish_manager.tree.root_item.children:
            if top_level_item.thumbnail_explicit:
                thumbnail_has_multiple_values = True

            if top_level_item.description not in [None, self._summary_comment]:
                description_had_multiple_values = True

            if description_had_multiple_values and thumbnail_has_multiple_values:
                break

            for descendant in top_level_item.descendants:
                if descendant.thumbnail_explicit:
                    # shortcut if one descendant has an explicit thumbnail
                    thumbnail_has_multiple_values = True

                if descendant.description not in [None, self._summary_comment]:
                    description_had_multiple_values = True

                if description_had_multiple_values and thumbnail_has_multiple_values:
                    break

            # break out of this loop if an explicit thumbnail and different description
            # was found for a descendant of the top level item.
            if thumbnail_has_multiple_values and description_had_multiple_values:
                break

        self.ui.item_thumbnail._set_multiple_values_indicator(
            thumbnail_has_multiple_values
        )
        self.ui.item_thumbnail.set_thumbnail(self._summary_thumbnail)

        # setting enabled to true to be able to take a snapshot to define the thumbnail
        self.ui.item_thumbnail.setEnabled(True)

        self.ui.item_description_label.setText("Description for all items")
        self.ui.item_inherited_item_label.hide()
        self.ui.item_comments.setText(self._summary_comment)
        self.ui.item_comments.setPlaceholderText("")

        # the item_comments PublishDescriptionFocus won't display placeholder text if it is in focus
        # so clearing the focus from that widget in order to see the <multiple values> warning once
        # the master summary details page is opened
        self.ui.item_comments.clearFocus()
        self.ui.item_comments._show_multiple_values = description_had_multiple_values

        # for the summary, attempt to display the appropriate context in the
        # context widget. if all publish items have the same context, display
        # that one. if there are multiple, show none and update the label to
        # reflect it.

        # iterate over all the tree items to find currently used contexts
        current_contexts = {}
        for it in QtGui.QTreeWidgetItemIterator(self.ui.items_tree):
            item = it.value()
            publish_instance = item.get_publish_instance()
            if isinstance(publish_instance, PublishItem):
                context = publish_instance.context
                context_key = str(context)
                current_contexts[context_key] = context

        if len(current_contexts) == 1:
            # only one context being used by current items. prepopulate it in
            # the summary view's context widget
            context_key = list(current_contexts.keys())[0]
            self.ui.context_widget.set_context(current_contexts[context_key])
            context_label_text = "Task and Entity Link to apply to all items:"
        else:
            self.ui.context_widget.set_context(
                None,
                task_display_override=" -- Multiple values -- ",
                link_display_override=" -- Multiple values -- ",
            )
            context_label_text = (
                "Currently publishing items to %s contexts. "
                "Override all items here:" % (len(current_contexts),)
            )

        self.ui.context_widget.show()
        self.ui.context_widget.enable_editing(True, context_label_text)

        # create summary for all items
        # no need to have a summary since the main label says summary
        self.ui.item_summary_label.hide()

        (num_items, summary) = self.ui.items_tree.get_full_summary()
        self.ui.item_summary.setText(summary)
        self.ui.item_type.setText("%d tasks to execute" % num_items)

    def _full_rebuild(self):
        """
        Full rebuild of the plugin state. Everything is recollected.
        """
        self._progress_handler.set_phase(self._progress_handler.PHASE_LOAD)
        self._progress_handler.push(
            "Collecting items to %s..." % (self._display_action_name)
        )

        previously_collected_files = self._publish_manager.collected_files
        self._publish_manager.tree.clear(clear_persistent=True)

        logger.debug("Refresh: Running collection on current session...")
        new_session_items = self._publish_manager.collect_session()

        logger.debug(
            "Refresh: Running collection on all previously collected external " "files"
        )
        new_file_items = self._publish_manager.collect_files(previously_collected_files)

        num_items_created = len(new_session_items) + len(new_file_items)
        num_errors = self._progress_handler.pop()

        if num_errors == 0 and num_items_created == 1:
            self._progress_handler.logger.info("One item discovered by publisher.")
        elif num_errors == 0 and num_items_created > 1:
            self._progress_handler.logger.info(
                "%d items discovered by publisher." % num_items_created
            )
        elif num_errors > 0:
            self._progress_handler.logger.error("Errors reported. See log for details.")

        # make sure the ui is up to date
        self._synchronize_tree()

        self._summary_comment = ""

        # select summary
        self.ui.items_tree.select_first_item()

        # reset the validation flag
        self._validation_run = False

    def _on_drop(self):
        """
        When someone drops stuff into the publish.
        """
        checked_item_paths = []
        for row in range(self.model.rowCount()):
            item = self.model.item(row)
            if item.checkState() == QtCore.Qt.Checked:
                for root, _, files in os.walk(item.text()):
                    if files:
                        if files[0].endswith('.exr') or files[0].endswith('.dpx'):
                            checked_item_paths.append(item.text())
                        else:
                            if 'temp' not in root:
                                if files[0].endswith('.mov'):
                                    for file in files:
                                        checked_item_paths.append(root + '\\' + file)
                            # else:
                            #     for file in files:
                            #         temp_file = root + '\\' + file
                            #         logger.info(f'test23:{temp_file}')

        # Short circuiting method disabling actual action performed on dropping to the target.
        if not self.manual_load_enabled:
            self._progress_handler.logger.error("Drag & drop disabled.")
            return

        # add files and rebuild tree
        self._progress_handler.set_phase(self._progress_handler.PHASE_LOAD)
        self._progress_handler.push("Processing external files...")

        # pyside gives us back unicode. Make sure we convert it to strings
        str_files = [six.ensure_str(f) for f in checked_item_paths]

        try:
            self.ui.main_stack.setCurrentIndex(self.PUBLISH_SCREEN)

            # ensure the progress details are parented here in case we get
            # stuck here.
            self._progress_handler.progress_details.set_parent(self.ui.main_frame)

            self._overlay.show_loading()
            self.ui.button_container.hide()
            new_items = self._publish_manager.collect_files(str_files)
            num_items_created = len(new_items)
            num_errors = self._progress_handler.pop()

            if num_errors == 0 and num_items_created == 0:
                self._progress_handler.logger.info("Nothing was added.")
            elif num_errors == 0 and num_items_created == 1:
                self._progress_handler.logger.info("One item was added.")
            elif num_errors == 0 and num_items_created > 1:
                self._progress_handler.logger.info(
                    "%d items were added." % num_items_created
                )
            elif num_errors == 1:
                self._progress_handler.logger.error(
                    "An error was reported. Please see the log for details."
                )
            else:
                self._progress_handler.logger.error(
                    "%d errors reported. Please see the log for details." % num_errors
                )

            # rebuild the tree
            self._synchronize_tree()

        finally:
            self._overlay.hide()
            self.ui.button_container.show()

        # lastly, select the summary
        self.ui.items_tree.select_first_item()

    def _synchronize_tree(self):
        """
        Redraws the ui and rebuilds data based on
        the low level plugin representation
        """
        top_level_items = list(self._publish_manager.tree.root_item.children)
        if len(top_level_items) == 0:
            if not self.manual_load_enabled:
                # No items collected and 'enable_manual_load' application option
                # false, display that special error overlay.
                self._show_no_items_error()
            else:
                # nothing in list. show the full screen drag and drop ui
                self.ui.main_stack.setCurrentIndex(self.DRAG_SCREEN)

        else:
            self.ui.main_stack.setCurrentIndex(self.PUBLISH_SCREEN)

            # ensure the progress details widget is available for overlay on the
            # main frame of the publish ui
            self._progress_handler.progress_details.set_parent(self.ui.main_frame)

            self.ui.items_tree.build_tree()

    def _set_tree_items_expanded(self, expanded):
        """
        Expand/Collapse all top-level publish items in the left side tree

        :param boole expanded: Expand if True, Collapse otherwise
        """

        for it in QtGui.QTreeWidgetItemIterator(self.ui.items_tree):
            item = it.value()
            if isinstance(item, TopLevelTreeNodeItem):
                item.setExpanded(expanded)

    def _set_description_inheritance_ui(self, tree_item):
        """
        Sets up the UI based on the selected item's description inheritance state.
        :param tree_item:
        :return:
        """
        self.ui.item_inherited_item_label.show()
        highlight_color = self._bundle.style_constants["SG_HIGHLIGHT_COLOR"]
        base_lbl = '<span style=" font-size:10pt;">{0}</span>'

        if isinstance(tree_item, TreeNodeItem) and tree_item.inherit_description:
            self.ui.item_comments.setStyleSheet(
                "border: 1px solid {0};".format(highlight_color)
            )

            lbl_prefix = base_lbl.format(
                'Description inherited from: <a href="inherited description" style="color: {0}">{1}</a>'
            )

            desc_item = self._find_inherited_description_item(tree_item)
            if desc_item:
                name = desc_item.get_publish_instance().name
                self.ui.item_inherited_item_label.setText(
                    lbl_prefix.format(highlight_color, name)
                )
                return
            elif self._summary_comment:
                # If the description is not inherited from another item, then it is inherited from the
                # summary
                self.ui.item_inherited_item_label.setText(
                    lbl_prefix.format(highlight_color, "Summary")
                )
                return

        self.ui.item_comments.setStyleSheet(
            "border: 0px solid {0};".format(highlight_color)
        )
        self.ui.item_inherited_item_label.setText(
            base_lbl.format("Description not inherited")
        )

    def _delete_selected(self):
        """
        Delete all selected items. Prompt the user.
        """
        # check with the user

        num_items = len(self.ui.items_tree.selectedItems())

        if num_items == 0:
            return

        if num_items > 1:
            msg = "This will remove %d items from the list." % num_items
        else:
            msg = "Remove the item from the list?"

        res = QtGui.QMessageBox.question(
            self, "Remove items?", msg, QtGui.QMessageBox.Ok | QtGui.QMessageBox.Cancel
        )

        if res == QtGui.QMessageBox.Cancel:
            return

        processing_items = []

        # delete from the tree
        for tree_item in self.ui.items_tree.selectedItems():
            if isinstance(tree_item, TreeNodeItem):
                processing_items.append(tree_item.item)

        for item in processing_items:
            self._publish_manager.tree.remove_item(item)

        self._synchronize_tree()

        self.ui.items_tree.select_first_item()

    def _check_all(self, checked):
        """
        Check all boxes in the currently active tree
        """

        def _check_r(parent):
            for child_index in range(parent.childCount()):
                child = parent.child(child_index)
                child.checkbox.setChecked(checked)
                _check_r(child)

        parent = self.ui.items_tree.invisibleRootItem()

        _check_r(parent)

    # set all nodes to "ready to go"
    def _reset_tree_icon_r(self, parent):
        """
        Clear the current progress icon recursively
        for the given tree node
        """
        for child_index in range(parent.childCount()):
            child = parent.child(child_index)
            child.reset_progress()
            self._reset_tree_icon_r(child)

    def _prepare_tree(self, number_phases):
        """
        Prepares the tree for processing.

        Will reset the progress bar and set it's max
        value based on the number of nodes plus the
        specified number of phases.

        Will clear status icons in the tree.

        :param int number_phases: Number of passes to run
        """

        self._set_tree_items_expanded(True)

        parent = self.ui.items_tree.invisibleRootItem()

        self._reset_tree_icon_r(parent)

        # set all nodes to "ready to go"
        def _begin_process_r(parent):
            total_number_nodes = 0
            for child_index in range(parent.childCount()):
                child = parent.child(child_index)
                if child.checked:
                    # child is ticked
                    total_number_nodes += 1
                total_number_nodes += _begin_process_r(child)
            return total_number_nodes

        total_number_nodes = _begin_process_r(parent)
        # reset progress bar
        self._progress_handler.reset_progress(total_number_nodes * number_phases)

    def do_validate(self, is_standalone=True):
        """
        Perform a full validation

        :param bool is_standalone: Indicates that the validation runs on its own,
            not part of a publish workflow.
        :returns: number of issues reported
        """

        # Make sure that settings from the current selection are retrieved from the UI and applied
        # back on the tasks.
        self._pull_settings_from_ui(self._current_tasks)

        if is_standalone:
            self._prepare_tree(number_phases=1)

        # inform the progress system of the current mode
        self._progress_handler.set_phase(self._progress_handler.PHASE_VALIDATE)
        self._progress_handler.push("Running validation pass")

        num_issues = 0
        self.ui.stop_processing.show()
        try:
            failed_to_validate = self._publish_manager.validate(
                task_generator=self._validate_task_generator(is_standalone)
            )
            num_issues = len(failed_to_validate)
        finally:
            self._progress_handler.pop()
            if self._stop_processing_flagged:
                self._progress_handler.logger.info("Processing aborted by user.")
            elif num_issues > 0:
                self._progress_handler.logger.error(
                    "Validation Complete. %d issues reported." % num_issues
                )
            else:
                self._progress_handler.logger.info(
                    "Validation Complete. All checks passed."
                )

            if is_standalone:
                # reset process aborted flag
                self._stop_processing_flagged = False
                self.ui.stop_processing.hide()
                # reset the progress
                self._progress_handler.reset_progress()

        # remember that validation has completed at least once
        self._validation_run = True

        return num_issues

    def do_publish(self):
        """
        Perform a full publish
        """
        publish_failed = False

        # hide the action buttons during publish
        self.ui.button_container.hide()

        # Make sure that settings from the current selection are retrieved from the UI and applied
        # back on the tasks.
        self._prepare_tree(number_phases=3)

        try:
            # show cancel button
            self.ui.stop_processing.show()

            # is the app configured to execute the validation when publish
            # is triggered?
            if self._bundle.get_setting("validate_on_publish"):
                # do_validate returns the number of issues encountered
                if self.do_validate(is_standalone=False) > 0:
                    self._progress_handler.logger.error(
                        "Validation errors detected. " "Not proceeding with publish."
                    )
                    self.ui.button_container.show()
                    return

            # validation not required on publish, it has already run though
            elif self._validation_run:
                self._progress_handler.logger.info(
                    "Skipping validation pass just before publish. "
                    "It was already run manually."
                )

            # validation not required on publish. no validation done yet
            else:
                # get user confirmation that they would like to continue
                button_clicked = QtGui.QMessageBox.question(
                    self,
                    "%s without Validation?" % (self._display_action_name,),
                    "You are attempting to %s without validation. Are you sure "
                    "you wish to continue?" % (self._display_action_name,),
                    buttons=QtGui.QMessageBox.Yes | QtGui.QMessageBox.Cancel,
                )
                if button_clicked == QtGui.QMessageBox.Cancel:
                    # user does not want ot continue.
                    self.ui.button_container.show()
                    return

                self._progress_handler.logger.info("User skipped validation step.")

            if self._stop_processing_flagged:
                # stop processing
                self.ui.button_container.show()
                return

            # inform the progress system of the current mode
            self._progress_handler.set_phase(self._progress_handler.PHASE_PUBLISH)
            self._progress_handler.push("Running publishing pass")

            # clear all icons
            parent = self.ui.items_tree.invisibleRootItem()
            self._reset_tree_icon_r(parent)

            try:
                self._publish_manager.publish(
                    task_generator=self._publish_task_generator(), CheckableItem = CheckableItem, colorspace = self._selected_color
                )
            except Exception:
                # ensure the full error shows up in the log file
                logger.error("Publish error stack:\n%s" % (traceback.format_exc(),))
                # and log to ui
                self._progress_handler.logger.error("Error while publishing. Aborting.")
                publish_failed = True
            finally:
                self._progress_handler.pop()

            if not publish_failed and not self._stop_processing_flagged:
                # proceed with finalizing phase

                # inform the progress system of the current mode
                self._progress_handler.set_phase(self._progress_handler.PHASE_FINALIZE)
                self._progress_handler.push("Running finalizing pass")

                try:
                    pass
                    # self._publish_manager.finalize(
                    #     task_generator=self._finalize_task_generator()
                    # )
                except Exception:
                    # ensure the full error shows up in the log file
                    logger.error(
                        "Finalize error stack:\n%s" % (traceback.format_exc(),)
                    )
                    # and log to ui
                    self._progress_handler.logger.error(
                        "Error while finalizing. Aborting."
                    )
                    publish_failed = True
                finally:
                    self._progress_handler.pop()

            # if stop processing was flagged, don't show summary at end
            if self._stop_processing_flagged:
                self._progress_handler.logger.info("Processing aborted by user.")
                self.ui.button_container.show()
                return

        finally:
            # hide cancel button
            self.ui.stop_processing.hide()
            # reset abort state
            self._stop_processing_flagged = False
            # reset the progress
            self._progress_handler.reset_progress()

        # disable validate and publish buttons
        # show close button instead
        self.ui.validate.hide()
        self.ui.publish.hide()
        self.ui.close.show()

        if publish_failed:
            self._progress_handler.logger.error(
                "Publish Failed! For details, click here."
            )
            self._overlay.show_fail()
        else:
            # Publish succeeded
            # Log the toolkit "Published" metric
            self._progress_handler.logger.info(
                "Publish Complete! For details, click here."
            )
            self._overlay.show_success()

    def _publish_again_clicked(self):
        """
        Slot that should be called when summary overlay close button is clicked.
        """

        # clear the tree and recollect the session
        self._publish_manager.tree.clear(clear_persistent=True)
        self._publish_manager.collect_session()

        self._synchronize_tree()

        # show publish and validate buttons
        self.ui.validate.show()
        self.ui.publish.show()
        self.ui.close.hide()

        self.ui.button_container.show()

        # hide summary overlay
        self._overlay.hide()

        # select summary
        self.ui.items_tree.select_first_item()

        # reset the validation flag
        self._validation_run = False

    def _find_inherited_description_item(self, item):
        """
        Recursively looks up the tree to find the item that the passed item
        should inherit the description from.
        If this item doesn't inherit then it will return None.
        It only considers parents, and so will not return the summary item even if
        that's what the description should ultimately inherit from.
        :return: TreeNodeItem or None
        """

        def _get_parent_item(item):
            if item is None:
                # We have reached the top of the tree without finding an item that doesn't inherit a description.
                return None
            elif not isinstance(item, TreeNodeItem) or item.inherit_description:
                return _get_parent_item(item.parent())
            else:
                # We've found an item that doesn't inherit its description so use this.
                return item

        return _get_parent_item(item.parent())

    def _find_inherited_description(self, node_item):
        """
        For a given item it will search recursively through it's parents to find a description.
        :param node_item: TreeNodeItem
        :return: str
        """
        inherited_desc_item = self._find_inherited_description_item(node_item)
        if inherited_desc_item:
            return inherited_desc_item.get_publish_instance().description
        else:
            # There is no parent item found in the tree that doesn't inherit, so we should grab the
            # description from the summary.
            return self._summary_comment

    def _get_tree_items(self):
        """
        Retrieves all the items from the tree.

        :returns: A list of QTreeItem.
        """
        # We used to be iterating on the items and yielding them
        # one after the other in the _task_generator method,
        # but that often gave us Internal C++ error about deleted objects.
        # It seems getting everything first and them returning a flat list
        # works.
        tree_iterator = QtGui.QTreeWidgetItemIterator(self.ui.items_tree)
        tree_items = []
        while tree_iterator.value():
            tree_items.append(tree_iterator.value())
            tree_iterator += 1

        return tree_items

    def _task_generator(self, stage_name):
        """
        This method yields tree items for our various stages. It will update the UI
        progress handler as items are visited.

        :param stage_name: Name of the current stage.
        """

        list_items = self._get_tree_items()

        for ui_item in list_items:

            if self._stop_processing_flagged:
                # jump out of the iteration
                break

            if not ui_item.checked:
                continue

            self._progress_handler.push(
                "%s: %s"
                % (
                    stage_name,
                    ui_item,
                ),
                ui_item.icon,
                ui_item.get_publish_instance(),
            )

            try:
                yield ui_item
            finally:
                self._progress_handler.increment_progress()
                self._progress_handler.pop()

    def _validate_task_generator(self, is_standalone):
        """
        This method generates tasks for the validation phase. It handles
        processing the tasks in the proper order and updating the task's UI
        representation in the dialog.
        """
        for ui_item in self._task_generator("Validating"):
            try:
                # yield each child item to be acted on by the publish api
                if isinstance(ui_item, TreeNodeTask):
                    (is_successful, error) = yield ui_item.task
                    error_msg = str(error)

                # all other nodes are UI-only and can handle their own
                # validation
                else:
                    is_successful = ui_item.validate(is_standalone)
                    error_msg = "Unknown validation error!"
            except Exception as e:
                ui_item.set_status_upwards(ui_item.STATUS_VALIDATION_ERROR, str(e))
                raise
            else:
                if is_successful:
                    if is_standalone:
                        ui_item.set_status(ui_item.STATUS_VALIDATION_STANDALONE)
                    else:
                        ui_item.set_status(ui_item.STATUS_VALIDATION)
                else:
                    ui_item.set_status_upwards(
                        ui_item.STATUS_VALIDATION_ERROR, error_msg
                    )

    def _publish_task_generator(self):
        """
        This method generates tasks for the publish phase. It handles
        processing the tasks in the proper order and updating the task's UI
        representation in the dialog.
        """

        for ui_item in self._task_generator("Publishing"):
            try:
                # yield each child item to be acted on by the publish api
                if isinstance(ui_item, TreeNodeTask):
                    yield ui_item.task

                # all other nodes are UI-only and can handle their own
                # publishing
                else:
                    ui_item.publish()
            except Exception as e:
                ui_item.set_status_upwards(ui_item.STATUS_PUBLISH_ERROR, str(e))
                raise
            else:
                ui_item.set_status(ui_item.STATUS_PUBLISH)

    def _finalize_task_generator(self):
        """
        This method generates tasks for the finalize phase. It handles
        processing the tasks in the proper order and updating the task's UI
        representation in the dialog.
        """
        for ui_item in self._task_generator("Finalizing"):
            try:
                # yield each child item to be acted on by the publish api
                if isinstance(ui_item, TreeNodeTask):
                    finalize_exception = yield ui_item.task

                # all other nodes are UI-only and can handle their own
                # finalize
                else:
                    ui_item.finalize()
                    finalize_exception = None
            except Exception as e:
                ui_item.set_status_upwards(ui_item.STATUS_FINALIZE_ERROR, str(e))
                raise
            else:
                if finalize_exception:
                    ui_item.set_status_upwards(
                        ui_item.STATUS_FINALIZE_ERROR, str(finalize_exception)
                    )
                else:
                    ui_item.set_status(ui_item.STATUS_FINALIZE)

    def _on_item_context_change(self, context):
        """
        Fires when a new context is selected for the current item
        """

        # For each of the scenarios below, we ensure that the item being updated
        # allows context change. The widget should be disabled for the single
        # item case, but we check to be completely sure. For the summary case,
        # we show the widget but we don't want to update selected items that
        # are on context change lockdown.
        sync_required = False

        # TODO: see todo below...
        # items_with_new_context = []

        if self._current_item is None:
            # this is the summary item - so update all items!
            for top_level_item in self._publish_manager.tree.root_item.children:
                if top_level_item.context_change_allowed:
                    top_level_item.context = context

                    # TODO: see todo below...
                    # this item and all of its descendents in the tree need to
                    # have their plugins reattached given the new context
                    # items_with_new_context.append(top_level_item)
                    # items_with_new_context.extend([i for i in top_level_item])

                    sync_required = True
        else:
            if self._current_item.context_change_allowed:
                self._current_item.context = context

                # TODO: see todo below...
                # this item and all of its descendents in the tree need to have
                # their plugins reattached given the new context
                # items_with_new_context.append(self._current_item)
                # items_with_new_context.extend([i for i in self._current_item])

                sync_required = True

        # TODO: attach plugins for the destination context. this is commented
        # out for now as it is a change in behavior and likely implies some
        # additional things to consider. for example: the drag/drop of items
        # within the tree (which changes context) will need to be updated. will
        # need to. also, will need to transfer task settings as the context
        # changes, where appropriate.
        #
        # ...
        # for any items that have a new context, rerun plugin attachment so that
        # they are published using the destination context's plugins
        # self._publish_manager.attach_plugins(items_with_new_context)

        if sync_required:
            self._synchronize_tree()

        # Make the task validation if the setting `task_required` exists and it's True
        self._validate_task_required()

    def _run_nuke_script(self, script_path):
        """
        This function runs the mode for placing Nuke files.

        :param script_path: File path to the Nuke script to run
        """

        config_path = r'X:\ShotGrid_Test_jw\Project\config_test'
        tk = sgtk.sgtk_from_path(config_path)
        sg = tk.shotgun

        current_engine = sgtk.platform.current_engine()
        context = current_engine.context
        user_id = context.user['id']

        if user_id == 319:
            nuke_executable = r"X:/Inhouse/Nuke15.1v2/Nuke15.1.exe"
        else:
            nuke_executable = r"X:/Inhouse/Nuke15.0v4/Nuke15.0.exe"

        checked_item_paths = self._get_checked_item_paths()

        for checked_item_path in checked_item_paths:
            parts = checked_item_path.split('/')
            origin_directory_path = parts[-1]
            project_code = parts[3]
            target_folder = None
            if os.path.isdir(checked_item_path):
                parts = parts[-1].split('_')

                # Extract the necessary parts
                # Drama
                if origin_directory_path.startswith("EP"):
                    seq_code = parts[0]
                    shot_code = parts[0] + '_' + parts[1] + '_' + parts[2]
                    category = parts[3]
                    version = parts[4]
                # Movie
                else:
                    seq_code = parts[0]
                    shot_code = parts[0] + '_' + parts[1]
                    category = parts[2]
                    version = parts[3]

                # Extract version information from source path
                target_folder = os.path.join(
                    r'X:\ShotGrid_Test_jw\Project',
                    project_code,
                    '04_SEQ',
                    seq_code,
                    shot_code,
                    'Plates',
                    category,
                    version
                )
                temp_folder_path = os.path.join(target_folder, 'temp')
                temp_folder_path = temp_folder_path.replace('\\', '/')
                new_output_path = os.path.join(temp_folder_path, f'{origin_directory_path}.mov')
                new_output_path = new_output_path.replace('\\', '/')

                if not os.path.isfile(new_output_path):
                    logger.info("Event occurred! Run the Nuke script.")
                    command = [nuke_executable, '-t', script_path]
                    self._run_subprocess(command)

                    for path in checked_item_paths:
                        self._determine_paths(path)

                    logger.info("Exit Nuke batch mode execution.")

    def _get_checked_item_paths(self):
        """
        Get the paths of the checked items in the model.
        """

        checked_item_paths = []
        path_metadatas = []

        for row in range(self.model.rowCount()):
            item = self.model.item(row)
            if item.checkState() == QtCore.Qt.Checked:
                for root, _, files in os.walk(item.text()):
                    if files:
                        if files[0].endswith('.exr') or files[0].endswith('.dpx'):
                            if 'temp' not in root:
                                metadata = self._get_metadata(root + '/' + files[0])
                                checked_item_paths.append(item.text())
                                path_metadatas.append(
                                    {'path': item.text(), 'metadata': metadata, 'colorspace': self._selected_color})
                        else:
                            if 'temp' not in root:
                                checked_item_paths.append(item.text())
                                for file in files:
                                    metadata = self._get_metadata(root + '/' + file)
                                    mov_path = os.path.join(root, file)
                                    mov_path = mov_path.replace('\\', '/')
                                    path_metadatas.append(
                                        {'path': mov_path, 'metadata': metadata, 'colorspace': self._selected_color})
        os.environ['TEMP_PATH'] = json.dumps(path_metadatas)
        return checked_item_paths

    def _get_metadata(self, file_path):
        if file_path.endswith('.exr'):
            return self._read_exr_metadata(file_path)
        elif file_path.endswith('.dpx'):
            return self._read_dpx_metadata(file_path)
        elif file_path.endswith('.mov'):
            return self._read_mov_metadata(file_path)

    @staticmethod
    def _read_exr_metadata(file_path):
        metadata = {}
        exr_file = InputFile(file_path)
        header = exr_file.header()
        metadata['width'] = str(header['dataWindow'].max.x - header['dataWindow'].min.x + 1)
        metadata['height'] = str(header['dataWindow'].max.y - header['dataWindow'].min.y + 1)
        timecode = header['timeCode']
        hours = timecode.hours
        minutes = timecode.minutes
        seconds = timecode.seconds
        frame = timecode.frame
        metadata['timecode'] = str(f"{hours:02}:{minutes:02}:{seconds:02}:{frame:02}")
        metadata['fps'] = str(header['framesPerSecond'])
        metadata['type'] = 'exr'

        return metadata

    @staticmethod
    def _read_mov_metadata(file_path):
        metadata = {}
        media_info = MediaInfo.parse(file_path)

        time_code_of_first_frame = None
        time_code_of_last_frame = None

        general_track = media_info.tracks[0]
        filename = general_track.file_name if hasattr(general_track, 'file_name') else "Unknown"
        duration = general_track.duration

        video_tracks = [track for track in media_info.tracks if track.track_type == 'Video']
        resolution = f"{video_tracks[0].width}x{video_tracks[0].height}" if video_tracks else "Unknown"
        width = video_tracks[0].width
        height = video_tracks[0].height

        fps = float(video_tracks[0].frame_rate) if video_tracks else "Unknown"

        for track in media_info.tracks:
            if track.track_type == 'Other':
                if hasattr(track, 'time_code_of_first_frame'):
                    time_code_of_first_frame = track.time_code_of_first_frame
                if hasattr(track, 'time_code_of_last_frame'):
                    time_code_of_last_frame = track.time_code_of_last_frame

        # Assuming the first frame is 1001
        first_frame = 1
        last_frame = int(duration / 1000 * fps)

        metadata['width'] = width
        metadata['height'] = height

        metadata['fps'] = fps
        metadata['first_frame'] = first_frame
        metadata['last_frame'] = last_frame
        metadata['type'] = 'mov'

        return metadata

    @staticmethod
    def _read_dpx_metadata(file_path):
        metadata = {}
        dpx_header = DpxHeaderEx(file_path)
        metadata['width'] = dpx_header.image_header.pixels_per_line
        metadata['height'] = dpx_header.image_header.lines_per_element
        metadata['timecode'] = dpx_header.tv_header.time_code
        fps_origin = dpx_header.film_header.frame_rate if dpx_header.film_header.frame_rate else dpx_header.tv_header.frame_rate
        metadata['fps'] = round(fps_origin, 3)
        metadata['type'] = 'dpx'

        return metadata

    @staticmethod
    def _run_subprocess(command):
        """
        Run a subprocess command and log its output.
        """
        try:
            process = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            for stdout_line in iter(process.stdout.readline, ""):
                logger.info(stdout_line.strip())
            for stderr_line in iter(process.stderr.readline, ""):
                logger.error(stderr_line.strip())
            process.stdout.close()
            process.stderr.close()
            return_code = process.wait()
            if return_code == 0:
                logger.info("Nuke script run successfully.")
            else:
                logger.error(f"Error running Nuke script. Return code: {return_code}")
        except subprocess.CalledProcessError as e:
            logger.error(f"An error occurred: {e}")

    def _determine_paths(self, path):
        """
        Determine the MOV and JPG paths based on the given path.
        """
        parts = path.split('/')
        origin_directory_path = parts[-1]
        project_code = parts[3]

        if os.path.isdir(path):
            parts = parts[-1].split('_')
        else:
            logger.error('Path does not exist')

        # Extract the necessary parts
        # Drama
        if origin_directory_path.startswith("EP"):
            seq_code = parts[0]
            shot_code = parts[0] + '_' + parts[1] + '_' + parts[2]
            category = parts[3]
            version = parts[4]
        # Movie
        else:
            seq_code = parts[0]
            shot_code = parts[0] + '_' + parts[1]
            category = parts[2]
            version = parts[3]

        # Extract version information from source path
        target_folder = os.path.join(
            r'X:\ShotGrid_Test_jw\Project',
            project_code,
            '04_SEQ',
            seq_code,
            shot_code,
            'Plates',
            category,
            version
        )

        temp_folder_path = target_folder + '/temp'

        if os.path.isdir(target_folder):
            files_in_folder = os.listdir(temp_folder_path)
            mov_files = [file for file in files_in_folder if file.endswith('.mov')]

            for mov_file in mov_files:
                mov_path = temp_folder_path + '/' + mov_file
                jpg_path = temp_folder_path + '/' + mov_file[:-4] + '.jpg'
                self._extract_frame_as_jpg(mov_path, jpg_path)

    @staticmethod
    def _check_file_exists(mov_path, jpg_path):
        """
        Check if the MOV and JPG files exist.
        """
        is_mov = os.path.isfile(mov_path)
        is_jpg = os.path.isfile(jpg_path)
        return is_mov and is_jpg

    @staticmethod
    def _extract_frame_as_jpg(mov_path, jpg_path):
        """
        Extract the first frame from a MOV file and save it as a JPG using FFmpeg.
        """
        if not os.path.isfile(jpg_path):
            ffmpeg_executable = r'X:\program\ffmpeg-master-latest-win64-gpl-shared\ffmpeg-master-latest-win64-gpl-shared\bin\ffmpeg.exe'
            logger.info(f"Extract the first frame from MOV file and save it as JPG: {mov_path} -> {jpg_path}")

            ffmpeg_command = [
                ffmpeg_executable,
                '-i', mov_path,
                '-vf', 'select=eq(n\,0)',
                '-q:v', '2',
                jpg_path
            ]

            try:
                logger.info(f"Run the FFmpeg command: {' '.join(ffmpeg_command)}")
                subprocess.run(ffmpeg_command, check=True, capture_output=True, text=True)
                logger.info("FFmpeg command execution completed.")
            except subprocess.CalledProcessError as e:
                logger.error(f"Error occurred while executing FFmpeg command: {e}")

    def _scan_listener(self):
        """
        A function that detects specific events.
        Here, we simulate an event with a conditional statement.
        """
        # Here, define the conditions for event occurrence
        event_occurred = True

        if event_occurred:
            self._run_nuke_script(self._default_config_path)
        else:
            logger.info("No event occurred.")

    def _copy_listener(self):
        """
        Copy the scanned files according to the toolkit folder structure.
        """
        checked_item_paths = self._get_checked_item_paths()
        for checked_item_path in checked_item_paths:
            if os.path.isdir(checked_item_path):
                temp_folder_name = 'temp'
                parts = checked_item_path.split('/')
                project_code = parts[3]
                origin_directory_path = parts[-1]
                parts = origin_directory_path.split('_')

                # Extract the necessary parts
                # Drama
                if origin_directory_path.startswith("EP"):
                    seq_code = parts[0]
                    shot_code = parts[0] + '_' + parts[1] + '_' + parts[2]
                    category = parts[3]
                    version = parts[4]
                # Movie
                else:
                    seq_code = parts[0]
                    shot_code = parts[0] + '_' + parts[1]
                    category = parts[2]
                    version = parts[3]

                # Extract version information from source path
                folder_name = os.path.basename(checked_item_path)
                if category in folder_name:
                    target_folder = os.path.join(
                        r'X:\ShotGrid_Test_jw\Project',
                        project_code,
                        '04_SEQ',
                        seq_code,
                        shot_code,
                        'Plates',
                        category,
                        version
                    )
                # Copy files to target folder
                if not os.path.isdir(target_folder):
                    self._copy_files(checked_item_path, target_folder, temp_folder_name, origin_directory_path)

    @staticmethod
    def _copy_files(source, target, temp_folder_name, origin_directory_path):
        # Start time record
        start_time = time.time()
        # Set file number starting value
        file_counter = 1001

        for root, dirs, files in os.walk(source):
            if temp_folder_name in dirs:
                dirs.remove(temp_folder_name)

            for file in files:
                # Full path to source file
                source_file = os.path.join(root, file)
                file_extension = os.path.splitext(file)[1]
                os.makedirs(target, exist_ok=True)
                new_filename = f"{origin_directory_path}{file_extension}" if file_extension == '.mov' else f"{origin_directory_path}.{file_counter:04d}{file_extension}"
                target_file = os.path.join(target, new_filename)

                # copy files
                shutil.copy2(source_file, target_file)
                logger.info(f'Copied {source_file} to {target_file}')

                file_counter += 1

        end_time = time.time()
        total_time = end_time - start_time
        logger.info(f'Total time taken: {total_time:.2f} seconds')

    def _on_browse(self, folders=False):
        """Opens a file dialog to browse to files for publishing."""

        # Redundant with disabling UI controls but short circuiting this method
        # further ensure that a user won't be able to browse for any file even
        # if a minor UI bug allows a way to do it.

        # if not self.manual_load_enabled:
        #     return

        # options for either browse type
        options = [
            QtGui.QFileDialog.DontResolveSymlinks,
            QtGui.QFileDialog.DontUseNativeDialog,
        ]

        if folders:
            # browse folders specifics
            caption = "Browse folders to publish image sequences"
            file_mode = QtGui.QFileDialog.Directory
            options.append(QtGui.QFileDialog.ShowDirsOnly)
        else:
            # browse files specifics
            caption = "Browse files to publish"
            file_mode = QtGui.QFileDialog.ExistingFiles

        # create the dialog
        file_dialog = QtGui.QFileDialog(parent=self, caption=caption)
        file_dialog.setLabelText(QtGui.QFileDialog.Accept, "Select")
        file_dialog.setLabelText(QtGui.QFileDialog.Reject, "Cancel")
        file_dialog.setFileMode(file_mode)
        file_dialog.setDirectory(self._default_directory_path)

        # set the appropriate options
        for option in options:
            file_dialog.setOption(option)

        # browse!
        if not file_dialog.exec_():
            return

        # process the browsed files/folders for publishing
        paths = file_dialog.selectedFiles()

        if paths:
            # simulate dropping the files into the dialog
            self.ui.lineEdit.setText(str(paths[0]))
            self._scan_path = str(paths[0])
            self._add_folder_to_list(str(paths[0]))

    def _add_folder_to_list(self, folder_path):
        subfolder_paths = [f.path for f in os.scandir(folder_path) if f.is_dir()]
        for idx, subfolder_path in enumerate(subfolder_paths):
            subfolder_path = subfolder_path.replace("\\", "/")
            subfolder_path_dict = {'index': idx, 'path': subfolder_path}
            self._add_file_to_list(subfolder_path_dict)

    def _add_file_to_list(self, file_path_dict):
        item = CheckableItem(file_path_dict, True)
        self.model.appendRow(item)

    def _get_jpeg_files(self, directory):
        """
        Get all .jpg files from the specified directory.

        :param directory: Path to the directory to search for .jpg files
        :return: List of paths to .jpg files in the directory
        """
        jpg_files = []
        for root, _, files in os.walk(directory):
            for file in files:
                if file.lower().endswith('.jpg'):
                    jpg_files.append(os.path.join(root, file))
        return jpg_files

    def _get_excel_data(self, date_name):
        bundle = sgtk.platform.current_bundle()
        default_directory = "X:/ShotGrid_Test_jw/Project/" + bundle.context.project.get("name", "Undefined") + '/scan' + '/' + date_name

        base_name = "scan_list_"
        ext = ".xlsx"

        # Find all files in the directory that match the base name and extension
        existing_files = [f for f in os.listdir(default_directory) if f.startswith(base_name) and f.endswith(ext)]

        if not existing_files:
            return None

        # Extract numbers from filenames and find the highest number
        numbers = [int(re.findall(r'\d+', f[len(base_name):])[0]) for f in existing_files if
                   re.findall(r'\d+', f[len(base_name):])]
        max_number = max(numbers) if numbers else 1
        next_file_name = f"{base_name}{max_number:02d}{ext}"

        excel_file_path = os.path.join(default_directory, next_file_name)

        if not excel_file_path:
            return None

        # Read the Excel file
        df = pd.read_excel(excel_file_path)

        # Filter checked and unchecked rows
        checked_df = df[df['Check'] == 'checked']
        unchecked_df = df[df['Check'] == 'unchecked']

        # Create a list of dictionaries containing index and row data for checked items
        checked_items_list = [{'index': idx, **row} for idx, row in checked_df.iterrows()]
        unchecked_items_list = [{'index': idx, **row} for idx, row in unchecked_df.iterrows()]
        all_items_list = checked_items_list + unchecked_items_list

        return all_items_list


    def _save_checked_items_to_excel(self):
        excel_item_list = []

        for row in range(self.model.rowCount()):
            item = self.model.item(row)
            excel_item = {}

            if item.checkState() == QtCore.Qt.Checked:
                excel_item['Check'] = 'checked'
            else:
                excel_item['Check'] = 'unchecked'

            checked_item_path = item.text()
            target_folder = None

            if os.path.isdir(checked_item_path):
                temp_folder_name = 'temp'
                parts = checked_item_path.split('/')
                origin_directory_path = parts[-1]
                project_code = parts[3]
                parts = parts[-1].split('_')

                # Extract the necessary parts
                # Drama
                if origin_directory_path.startswith("EP"):
                    seq_code = parts[0]
                    shot_code = parts[0] + '_' + parts[1] + '_' + parts[2]
                    category = parts[3]
                    version = parts[4]
                # Movie
                else:
                    seq_code = parts[0]
                    shot_code = parts[0] + '_' + parts[1]
                    category = parts[2]
                    version = parts[3]

                # Extract version information from source path
                target_folder = os.path.join(
                    r'X:\ShotGrid_Test_jw\Project',
                    project_code,
                    '04_SEQ',
                    seq_code,
                    shot_code,
                    'Plates',
                    category,
                    version
                )

            temp_path = target_folder + '/temp'
            jpg_files = self._get_jpeg_files(temp_path)

            if not jpg_files:
                jpg_file = ''
            else:
                jpg_file = jpg_files[0].replace('\\', '/')

            file_list = os.listdir(item.text())

            ## mov 파일은 폴더당 하나라고 가정
            files = [f for f in file_list if f.endswith('.exr') or f.endswith('dpx') or f.endswith('mov')]
            first_file = files[0]

            metadata = self._get_metadata(item.text() + '/' + first_file)

            path_segments = item.text().split("/")
            folder_name = path_segments[6]
            date_name = path_segments[5]
            parts = folder_name.split("_")
            sequence_code = parts[0]

            # Drama
            if origin_directory_path.startswith("EP"):
                shot_code = "_".join(parts[0:3])
            else:
                shot_code = "_".join(parts[0:2])

            excel_item['Thumbnail'] = jpg_file
            excel_item['Sequence'] = sequence_code or ''
            excel_item['Shot'] = shot_code or ''
            excel_item['Type'] = category
            excel_item['Version'] = version

            if first_file.endswith('mov'):
                first_frame = metadata['first_frame'] + 1000
                last_frame = metadata['last_frame'] + 1000
                base_filename = first_file.rsplit('.', 1)[0]
                excel_item['Org_Clip'] = str(base_filename)
            else:
                match = re.search(r'(.*?)(\d+)\.(dpx|exr)$', first_file)
                prefix = match.group(1)
                number_str = match.group(2)
                suffix = match.group(3)
                first_frame = 1001
                last_frame = first_frame + len(files) - 1
                excel_item['Org_Clip'] = str(prefix[:-1])

            fps_number = metadata['fps']
            if metadata['type'] == 'exr':
                fps_value = metadata['fps']
                fps_number = re.search(r'\((.*?)\)', fps_value).group(1)
            excel_item['Cut_In'] = str(first_frame) or 0
            excel_item['Cut_Out'] = str(last_frame) or 0
            duration = last_frame - first_frame + 1
            excel_item['Duration'] = str(duration)

            width = metadata['width']
            height = metadata['height']
            excel_item['Resolution'] = str(f'{width}X{height}') or ''
            if str(fps_number) == 'nan':
                fps_number = '23.976'
            excel_item['Fps'] = str(fps_number)
            excel_item_list.append(excel_item)

        if excel_item_list:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Items"

            # Add header
            headers = ["Thumbnail", "Sequence", "Shot_Code", "Type", "Cut_In", "Cut_Out", "Duration", "Org_Clip",
                       "Resolution", "Fps", "Version", "Check"]
            gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

            for col, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.fill = gray_fill
                cell.border = border
                cell.font = Font(bold=True)

            checked_items = self._get_excel_data(date_name)

            for idx, item in enumerate(excel_item_list, start=2):
                thumbnail_path = item.get('Thumbnail')
                sequence = item.get('Sequence')
                shot = item.get('Shot')
                type = item.get('Type')
                check = item.get('Check')
                cut_in = item.get('Cut_In')
                cut_out = item.get('Cut_Out')
                duration = item.get('Duration')
                original_clip = item.get('Org_Clip')
                resolution = item.get('Resolution')
                fps = item.get('Fps')
                version = item.get('Version')

                if checked_items:
                    for checked_item in checked_items:
                        if checked_item["Shot_Code"] + '_' + checked_item["Type"] + '_' + checked_item["Version"] == shot + '_' + type + '_' + version:
                            if checked_item["Check"] == 'checked':
                                check = checked_item["Check"]

                if thumbnail_path and os.path.isfile(thumbnail_path):
                    img = Image(thumbnail_path)
                    img.anchor = f'A{idx}'  # Set the cell where the image will be placed
                    img.width = 150
                    img.height = 100
                    sheet.add_image(img)

                    # Adjust column width to fit the image
                    sheet.row_dimensions[idx].height = 75  # OpenPyXL row height units are different
                    col_letter = get_column_letter(1)  # 'A' column for Thumbnail
                    sheet.column_dimensions[col_letter].width = 18.75

                else:
                    thumbnail_cell = sheet.cell(row=idx, column=1, value="No thumbnail")
                    thumbnail_cell.border = border

                sequence_cell = sheet.cell(row=idx, column=2, value=sequence)
                sequence_cell.border = border
                shot_cell = sheet.cell(row=idx, column=3, value=shot)
                shot_cell.border = border
                type_cell = sheet.cell(row=idx, column=4, value=type)
                type_cell.border = border
                cut_in_cell = sheet.cell(row=idx, column=5, value=cut_in)
                cut_in_cell.border = border
                cut_out_cell = sheet.cell(row=idx, column=6, value=cut_out)
                cut_out_cell.border = border
                duration_cell = sheet.cell(row=idx, column=7, value=duration)
                duration_cell.border = border
                original_clip_cell = sheet.cell(row=idx, column=8, value=original_clip)
                original_clip_cell.border = border
                resolution_cell = sheet.cell(row=idx, column=9, value=resolution)
                resolution_cell.border = border
                fps_cell = sheet.cell(row=idx, column=10, value=fps)
                fps_cell.border = border
                version_cell = sheet.cell(row=idx, column=11, value=version)
                version_cell.border = border
                check_cell = sheet.cell(row=idx, column=12, value=check)
                check_cell.border = border

                # path_cell = sheet.cell(row=idx, column=7, value=item_path)
                # path_cell.border = border

            # Adjust column width to fit the contents
            for col in range(2, sheet.max_column + 1):
                max_length = 0
                column = get_column_letter(col)
                for cell in sheet[column]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width

            save_path = self._get_save_path(self._default_directory_path + '/' + date_name)
            default_directory = os.path.dirname(save_path)
            default_file_name = os.path.basename(save_path)
            options = QtGui.QFileDialog.Options()

            file_path, _ = QtGui.QFileDialog.getSaveFileName(
                self,
                "Save File",
                os.path.join(default_directory, default_file_name),
                "Excel Files (*.xlsx);;All Files (*)",
                options=options
            )

            if file_path:
                workbook.save(save_path)

    @staticmethod
    def _get_save_path(folder_path):
        base_name = "scan_list_"
        ext = ".xlsx"
        existing_files = [f for f in os.listdir(folder_path) if f.startswith(base_name) and f.endswith(ext)]

        if not existing_files:
            return os.path.join(folder_path, f"{base_name}01{ext}")

        numbers = [int(re.findall(r'\d+', f[len(base_name):])[0]) for f in existing_files if
                   re.findall(r'\d+', f[len(base_name):])]
        next_number = max(numbers) + 1 if numbers else 1
        next_file_name = f"{base_name}{next_number:02d}{ext}"
        return os.path.join(folder_path, next_file_name)

    def _on_excel_browse(self):
        parts = self._scan_path.split('/')
        date_name = parts[-1]
        project_name = "%s" % self._bundle.context.project.get("name", "Undefined")
        default_open_path = "X:/ShotGrid_Test_jw/Project/" + project_name + '/scan' + '/' + date_name
        options = QtGui.QFileDialog.Options()
        file_path, _ = QtGui.QFileDialog.getOpenFileName(
            self,
            "Select Excel File",
            default_open_path,
            "Excel Files (*.xlsx);;All Files (*)",
            options=options
        )

        if file_path:
            self._open_excel_file(file_path)

    @staticmethod
    def _open_excel_file(file_path):
        try:
            os.startfile(file_path)
        except Exception as e:
            logger.info(f'An error occurred while opening the file: {e}')

    def _check_all_items(self):
        for row in range(self.model.rowCount()):
            item = self.model.item(row)
            item.setCheckState(QtCore.Qt.Checked)

    def _uncheck_all_items(self):
        for row in range(self.model.rowCount()):
            item = self.model.item(row)
            item.setCheckState(QtCore.Qt.Unchecked)

    def _open_url(self, url):
        """Opens the supplied url in the appropriate browser."""
        try:
            logger.debug("Opening url: '%s'." % (url,))
            QtGui.QDesktopServices.openUrl(QtCore.QUrl(url))
        except Exception as e:
            logger.error("Failed to open url: '%s'. Reason: %s" % (url, e))

    def _trigger_stop_processing(self):
        """
        Triggers a request to stop processing
        """
        logger.info("Processing aborted.")
        self._stop_processing_flagged = True

    def _show_no_items_error(self):
        """
        Re-organize the UI for presenting the overlay with a special error message
        when the 'enable_manual_load' application option is false and there is no
        items collected.
        """
        # Hide everything but the close button.
        self.ui.validate.hide()
        self.ui.publish.hide()
        self.ui.button_container.hide()
        self.ui.progress_bar.hide()
        self.ui.close.show()

        self._progress_handler.logger.error("Drag & drop disabled.")

        self.ui.main_stack.setCurrentIndex(self.PUBLISH_SCREEN)
        self._overlay.show_no_items_error()

    def _validate_task_required(self):
        """
        Validates that a task is selected for every item and disables/enables the
        validate and publish buttons and finally change the color for the task
        label.
        """
        # Avoid validation if the setting `task_required` is False or not exists
        if not self._bundle.get_setting("task_required"):
            return

        all_items_selected_task = True
        for context_index in range(self.ui.items_tree.topLevelItemCount()):
            context_item = self.ui.items_tree.topLevelItem(context_index)

            if hasattr(context_item, "context") and not context_item.context.task:
                all_items_selected_task = False
                break

        if not all_items_selected_task:
            # disable buttons
            self.ui.publish.setEnabled(False)
            self.ui.validate.setEnabled(False)
            # change task label color to RED
            self.ui.context_widget.ui.task_label.setStyleSheet("color: red")
        else:
            # enable buttons
            self.ui.publish.setEnabled(True)
            self.ui.validate.setEnabled(True)

            # change task label color to the default value
            self.ui.context_widget.ui.task_label.setStyleSheet("")


class _TaskSelection(object):
    """
    Allows to manipulate a task selection as if it was a single object. It will hold a list of
    publish tasks from the tree items. It also removes the tedium of testing for an empty array
    and indexing [0] when doing comparisons.

    This class assumes that every task is of the same plugin type.

    :param items: List of task for in the selection. Defaults to an empty list.
    """

    def __init__(self, items=None):
        self._items = items or []

    def is_same_task_type(self, task_selection):
        """
        Checks if this task selection is of the same type as another task selection.

        :param task_selection: List of :class:`Task`.

        :returns: ``True`` is the plugins are the same, ``False`` otherwise. Note that
            two empty lists are considered to be of the same type.
        """
        if self._items and task_selection._items:
            # we know all the items in the list are of the same type, so we can
            # simply compare them using the first item of each list.
            return self._items[0].is_same_task_type(task_selection._items[0])
        elif not self._items and not task_selection._items:
            return True
        else:
            return False

    @property
    def has_custom_ui(self):
        """
        Checks if this selection has a custom UI.

        :returns: ``True`` if the selection uses a custom UI, ``False`` otherwise.
        """
        if self._items:
            return self._items[0].plugin.has_custom_ui
        else:
            return False

    @property
    def plugin(self):
        """
        Returns the plugin associated with this selection.

        :returns: The :class:`Plugin` instance or ``None``.
        """
        if self._items:
            return self._items[0].plugin
        else:
            return None

    def get_settings(self, widget):
        """
        Retrieves the settings from the selection's custom UI.

        :param widget: Custom UI's widget.

        :returns: Dictionary of settings as regular Python literals.
        """
        if self._items:
            # Get the publish items associated with the selected tasks.
            publish_items = self.get_task_items()
            return self._items[0].plugin.run_get_ui_settings(widget, publish_items)
        else:
            return {}

    def set_settings(self, widget, settings):
        """
        Sets the settings from the selection into the custom UI.

        :param widget: Custom UI's widget.
        :param settings: List of settings for all tasks.
        """
        if self._items:
            # Get the publish items associated with the selected tasks.
            publish_items = self.get_task_items()
            self._items[0].plugin.run_set_ui_settings(widget, settings, publish_items)

    def get_task_items(self):
        """
        Gets a list of items that the selected tasks are parented too.
        :return: list of PublishItems.
        """
        return [task.item for task in self._items]

    def __iter__(self):
        """
        Allows to iterate over items in the selection.
        """
        return iter(self._items)

    def __eq__(self, other):
        """
        Tests two selections for equality.
        """
        return self._items == other._items

    def __bool__(self):
        """
        :returns: ``True`` is the selection is not empty, ``False`` otherwise.
        """
        return bool(self._items)

    # To maintain Python 2 compatibility, define __nonzero__ as well as __bool__
    __nonzero__ = __bool__
